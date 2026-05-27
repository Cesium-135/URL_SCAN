#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
URL Jump Chain Capture & Classification Tool (Multithreaded + Click Simulation + Breakpoint Resume)
URL跳转链路捕获与分类工具（多线程 + 点击模拟 + 断点续传）

Before running, install dependencies:
    pip install playwright openpyxl matplotlib
    playwright install chromium

Usage:
    python script.py -i input.xlsx -o output.xlsx
"""

# -------------------------- 1. Imports | 导入库 --------------------------
import os
import re
import argparse
import logging
import traceback
from datetime import datetime
from urllib.parse import urlparse, urlunparse
from concurrent.futures import ThreadPoolExecutor, as_completed

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
import matplotlib.pyplot as plt

# -------------------------- 2. Global Configuration | 全局配置 --------------------------
CONFIG = {
    # File & Sheet Names | 文件与工作表名称
    "input_excel_path": "",          # Will be set by CLI or interactive
    "output_excel_path": "",         # Will be set by CLI or interactive
    "input_sheet_name": "Feuil1",    # Sheet name containing URLs
    "summary_sheet_name": "summary", # Summary sheet name
    "detail_sheet_name": "jump_chain", # Detail sheet name
    "url_column": "A",               # Excel column containing URLs (e.g., "A", "B") | URL所在列

    # Browser & Network | 浏览器与网络
    "headless_mode": True,           # Headless mode for cloud/server
    "timeout": 30000,                # Page timeout (ms)
    "max_retry_times": 2,            # Retry count per URL
    "max_jump_count": 20,            # Max redirect steps per URL
    "wait_for_network_idle": 2000,   # ms to wait after page load
    "proxy_config": None,            # Proxy dict: {"server": "http://ip:port"}

    # Multi-threading | 多线程
    "max_workers": 30,               # Number of concurrent threads

    # Click Simulation | 点击模拟
    "click_config": {
        "enable_click_function": True,
        "max_click_retry": 2,
        "wait_after_click": 5000,
        "capture_new_window": True,
        "target_keywords": {
            "auth_login": ["login", "signin", "sign-in", "secure access", "account",
                           "connexion", "se connecter", "accès sécurisé", "compte"],
            "espace_access": ["espace", "my espace", "access espace", "enter", "access",
                              "mon espace", "accéder à l'espace", "entrer", "accéder"]
        },
        "selector_priority": [
            "a[href*={kw}], button[id*={kw}], button[name*={kw}], a[class*={kw}]",
            "button:has-text('{kw}'), a:has-text('{kw}'), div[role='button']:has-text('{kw}')",
            "[role='button']:has-text('{kw}'), [role='link']:has-text('{kw}')"
        ]
    },

    # Cookie Consent Handling | Cookie 同意弹窗处理
    "cookie_config": {
        "enable_cookie_handle": True,
        "keywords": ["accept cookie", "accept all cookies", "accepter les cookies",
                     "tout accepter", "ok", "agree", "allow all"],
        "selectors": [
            "button:has-text('{kw}'), a:has-text('{kw}')",
            "[role='button']:has-text('{kw}'), div[class*='cookie'] button"
        ]
    },

    # URL Classification Rules (EN/FR only) | 分类规则
    "classify_rules": {
        "Authentication": {
            "keywords": ["login", "auth", "iam", "sso", "token", "signin", "authentification"],
            "page_features": ["username", "password", "login", "sign in",
                              "nom d'utilisateur", "mot de passe", "connexion"],
            "sub_class": {
                "SSO/Auth Service": ["sso", "iam", "oauth", "openid", "authentification unique"],
                "Login Page": ["login", "signin", "connexion", "se connecter"],
                "Token/Credential": ["token", "credential", "jwt", "bearer", "jeton"]
            }
        },
        "Espace": {
            "keywords": ["espace", "espresso", "station", "business", "portal"],
            "page_features": ["login required", "please log in",
                              "connexion requise", "veuillez vous connecter"],
            "sub_class": {
                "Station Espace": ["station", "espresso station", "station espresso"],
                "Business Espace": ["business", "entreprise", "professionnel"],
                "User Espace": ["mon espace", "my espace", "espace utilisateur"]
            }
        },
        "Backend/API": {
            "keywords": ["api", "gateway", "backend", "server", "interface", "rest"],
            "request_type": ["xhr", "fetch"],
            "sub_class": {
                "REST API": ["rest", "api/v1", "api/v2", "json", "xml"],
                "Gateway API": ["gateway", "proxy", "api gateway", "passerelle api"],
                "Internal API": ["internal", "backend", "serveur interne"]
            }
        },
        "Public front": {
            "default": True,
            "sub_class": {
                "Landing Page": ["home", "accueil", "landing", "welcome"],
                "Product Page": ["product", "produit", "service", "offre"],
                "Help/Support": ["help", "aide", "support", "assistance", "faq"]
            }
        },
        "Error": {
            "sub_class": {
                "Connection Error": ["ERR_CONNECTION", "connection refused", "connexion refusée"],
                "Timeout Error": ["ERR_TIMED_OUT", "timeout", "délai expiré"],
                "Certificate Error": ["ERR_CERT", "certificate", "certificat invalide"],
                "4XX/5XX HTTP": ["404", "403", "500", "502", "http error", "erreur http"],
                "Auth Required": ["401", "unauthorized", "non autorisé", "authentification requise"]
            }
        }
    },

    # Logging | 日志
    "log_config": {
        "log_file": "url_scan_{}.log".format(datetime.now().strftime("%Y%m%d_%H%M%S")),
        "log_level": "INFO"
    }
}

# -------------------------- 3. Logging Setup | 日志配置 --------------------------
logging.basicConfig(
    level=getattr(logging, CONFIG["log_config"]["log_level"]),
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(CONFIG["log_config"]["log_file"], encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# -------------------------- 4. Utility Functions | 工具函数 --------------------------
def parse_command_line_args():
    """Parse CLI arguments or prompt interactively"""
    parser = argparse.ArgumentParser(description="URL Scan Tool | URL扫描工具")
    parser.add_argument("-i", "--input", help="Input Excel file path | 输入Excel文件路径")
    parser.add_argument("-o", "--output", help="Output Excel file path | 输出Excel文件路径")
    args = parser.parse_args()

    if not args.input:
        args.input = input("Please enter input Excel file path: ").strip()
        while not args.input:
            args.input = input("Input path cannot be empty, please re-enter: ").strip()

    if not args.output:
        default_output = f"URL_SCAN_RESULT_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        args.output = input(f"Please enter output Excel path (default: {default_output}): ").strip() or default_output

    CONFIG["input_excel_path"] = args.input
    CONFIG["output_excel_path"] = args.output
    logger.info(f"Input: {CONFIG['input_excel_path']} | Output: {CONFIG['output_excel_path']}")

def standardize_url(url: str) -> str:
    """Standardize URL, add https if missing"""
    url = url.strip()
    if not url:
        return ""
    parsed = urlparse(url)
    if not parsed.scheme:
        url = "https://" + url
        parsed = urlparse(url)
    if not parsed.netloc or parsed.scheme not in ["http", "https"]:
        return ""
    return urlunparse(parsed).rstrip("/")

def is_valid_url(url: str) -> bool:
    return bool(url and urlparse(url).scheme in ["http", "https"] and urlparse(url).netloc)

def extract_error_code(error_msg: str) -> tuple:
    """Extract error code and error type from error message"""
    if not error_msg:
        return "", "Unknown Error"
    # Match Playwright/Chrome error codes
    error_code = re.search(r"ERR_\w+", error_msg)
    error_code = error_code.group() if error_code else ""
    if not error_code:
        http_code = re.search(r"\b(4\d{2}|5\d{2}|3\d{2})\b", error_msg)
        error_code = http_code.group() if http_code else ""
    # Classify
    error_type = "Unknown Error"
    for type_name, keywords in CONFIG["classify_rules"]["Error"]["sub_class"].items():
        if any(kw in error_msg for kw in keywords) or (error_code and any(kw in error_code for kw in keywords)):
            error_type = type_name
            break
    return error_code, error_type

def classify_url(jump_info: dict) -> tuple:
    """Return (main_class, sub_class) based on URL and page content"""
    if jump_info.get("error_msg"):
        _, error_type = extract_error_code(jump_info["error_msg"])
        return "Error", error_type

    url_lower = jump_info.get("jump_url", "").lower()
    page_html = jump_info.get("page_html", "").lower()
    request_type = jump_info.get("request_type", "")

    # Authentication
    auth_rule = CONFIG["classify_rules"]["Authentication"]
    if any(kw in url_lower for kw in auth_rule["keywords"]) or \
       any(feature in page_html for feature in auth_rule["page_features"]):
        sub = "Other Authentication"
        for sub_name, sub_kws in auth_rule["sub_class"].items():
            if any(kw in url_lower for kw in sub_kws):
                sub = sub_name
                break
        return "Authentication", sub

    # Espace
    espace_rule = CONFIG["classify_rules"]["Espace"]
    if any(kw in url_lower for kw in espace_rule["keywords"]) or \
       any(feature in page_html for feature in espace_rule["page_features"]):
        sub = "Other Espace"
        for sub_name, sub_kws in espace_rule["sub_class"].items():
            if any(kw in url_lower for kw in sub_kws):
                sub = sub_name
                break
        return "Espace", sub

    # Backend/API
    api_rule = CONFIG["classify_rules"]["Backend/API"]
    if any(kw in url_lower for kw in api_rule["keywords"]) or \
       request_type in api_rule["request_type"]:
        sub = "Other API"
        for sub_name, sub_kws in api_rule["sub_class"].items():
            if any(kw in url_lower for kw in sub_kws):
                sub = sub_name
                break
        return "Backend/API", sub

    # Public front
    sub = "Other Public Page"
    for sub_name, sub_kws in CONFIG["classify_rules"]["Public front"]["sub_class"].items():
        if any(kw in url_lower for kw in sub_kws):
            sub = sub_name
            break
    return "Public front", sub

def handle_cookie_consent(page):
    """Accept cookie popups if present"""
    if not CONFIG["cookie_config"]["enable_cookie_handle"]:
        return
    cfg = CONFIG["cookie_config"]
    for kw in cfg["keywords"]:
        for selector_tpl in cfg["selectors"]:
            selector = selector_tpl.format(kw=kw.lower())
            try:
                # Try main frame
                element = page.locator(selector).first
                if element.is_visible(timeout=1000) and element.is_enabled():
                    element.click()
                    logger.info(f"Clicked cookie consent: {selector}")
                    page.wait_for_timeout(1000)
                    return
            except:
                pass
    # Try iframes
    try:
        for frame in page.frames:
            for kw in cfg["keywords"]:
                for selector_tpl in cfg["selectors"]:
                    selector = selector_tpl.format(kw=kw.lower())
                    try:
                        element = frame.locator(selector).first
                        if element.is_visible(timeout=1000) and element.is_enabled():
                            element.click()
                            logger.info(f"Clicked cookie in iframe: {selector}")
                            page.wait_for_timeout(1000)
                            return
                    except:
                        pass
    except:
        pass

def find_clickable_element(page):
    """Find element to click for Espace/Login"""
    click_cfg = CONFIG["click_config"]
    all_keywords = click_cfg["target_keywords"]["auth_login"] + click_cfg["target_keywords"]["espace_access"]
    all_keywords = list(set([kw.lower() for kw in all_keywords]))
    for selector_tpl in click_cfg["selector_priority"]:
        for kw in all_keywords:
            selector = selector_tpl.format(kw=kw)
            try:
                element = page.locator(selector).first
                if element.is_visible(timeout=1000) and element.is_enabled():
                    return element, f"Selector: {selector}, Keyword: {kw}"
            except:
                continue
    return None, "No matching element"

# -------------------------- 5. Core Scanning Function (per URL) --------------------------
def scan_single_url(raw_url: str) -> dict:
    """Scan a single URL, capture jump chain, classification, errors."""
    result = {
        "raw_url": raw_url,
        "standard_url": "",
        "jump_chain": [],
        "http_status_list": [],
        "redirect_type_list": [],
        "trigger_mode_list": [],
        "main_class": "",
        "sub_class": "",
        "error_msg": "",
        "error_code": "",
        "auth_required": False,
        "click_triggered": False,
        "click_element_info": "",
        "final_url": "",
        "scan_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "processing_status": "Failed"
    }

    std_url = standardize_url(raw_url)
    result["standard_url"] = std_url
    if not is_valid_url(std_url):
        result["error_msg"] = "Invalid URL format"
        result["error_code"] = "INVALID_URL"
        result["main_class"], result["sub_class"] = classify_url({"error_msg": result["error_msg"]})
        result["processing_status"] = "Failed"
        logger.error(f"Invalid URL: {raw_url}")
        return result

    retry = 0
    while retry <= CONFIG["max_retry_times"]:
        playwright = None
        browser = None
        context = None
        page = None
        new_page_obj = None
        try:
            playwright = sync_playwright().start()
            browser_kwargs = {"headless": CONFIG["headless_mode"]}
            if CONFIG["proxy_config"]:
                browser_kwargs["proxy"] = CONFIG["proxy_config"]
            browser = playwright.chromium.launch(**browser_kwargs)
            context = browser.new_context(ignore_https_errors=True)
            page = context.new_page()
            page.set_default_timeout(CONFIG["timeout"])

            # Event storage
            navigation_events = []
            jump_id = 1
            current_trigger = "Auto Redirect"

            # Listeners
            def on_frame_navigated(frame):
                nonlocal jump_id, current_trigger
                if frame == page.main_frame:
                    url = frame.url
                    if not navigation_events or navigation_events[-1]["jump_url"] != url:
                        navigation_events.append({
                            "jump_id": jump_id,
                            "jump_url": url,
                            "from_url": page.url if jump_id > 1 else std_url,
                            "jump_type": "Client Navigation",
                            "status_code": 200,
                            "trigger_type": current_trigger,
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                        })
                        jump_id += 1

            def on_response(response):
                nonlocal jump_id, current_trigger
                if 300 <= response.status < 400:
                    location = response.headers.get("location", "")
                    if location:
                        redir_url = standardize_url(location)
                        if redir_url:
                            navigation_events.append({
                                "jump_id": jump_id,
                                "jump_url": redir_url,
                                "from_url": response.request.url,
                                "jump_type": "3xx Redirect",
                                "status_code": response.status,
                                "trigger_type": current_trigger,
                                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                            })
                            jump_id += 1

            def on_page_opened(new_page):
                nonlocal new_page_obj
                new_page_obj = new_page
                logger.info(f"New window opened: {new_page.url}")

            page.on("framenavigated", on_frame_navigated)
            page.on("response", on_response)
            if CONFIG["click_config"]["capture_new_window"]:
                context.on("page", on_page_opened)

            # Step 1: Go to URL
            response = page.goto(std_url, wait_until="networkidle", timeout=CONFIG["timeout"])
            page.wait_for_timeout(CONFIG["wait_for_network_idle"])
            handle_cookie_consent(page)

            # Check authentication required
            try:
                page_html = page.content()
            except Exception as e:
                logger.warning(f"Failed to get page content for auth check: {e}")
                page_html = ""
            auth_indicators = ["authentication required", "certificate required", "nom d'utilisateur",
                               "mot de passe", "authentification requise", "certificatif requis"]
            if any(ind in page_html.lower() for ind in auth_indicators):
                result["auth_required"] = True
                logger.info(f"Auth/certificate required: {std_url}")

            # Step 2: Click simulation
            if CONFIG["click_config"]["enable_click_function"]:
                element, desc = find_clickable_element(page)
                if element:
                    result["click_triggered"] = True
                    result["click_element_info"] = desc
                    for _ in range(CONFIG["click_config"]["max_click_retry"]):
                        try:
                            element.scroll_into_view_if_needed()
                            element.click()
                            current_trigger = "Manual Click"
                            page.wait_for_timeout(CONFIG["click_config"]["wait_after_click"])
                            if new_page_obj and CONFIG["click_config"]["capture_new_window"]:
                                # Wait for new page load
                                new_page_obj.wait_for_load_state("networkidle", timeout=CONFIG["timeout"])
                                final_new_url = new_page_obj.url
                                if is_valid_url(final_new_url):
                                    navigation_events.append({
                                        "jump_id": jump_id,
                                        "jump_url": final_new_url,
                                        "from_url": page.url,
                                        "jump_type": "New Window",
                                        "status_code": 200,
                                        "trigger_type": current_trigger,
                                        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                                    })
                                    jump_id += 1
                            break
                        except Exception as e:
                            logger.warning(f"Click retry failed: {e}")

            # Step 3: Determine final URL
            if new_page_obj and CONFIG["click_config"]["capture_new_window"]:
                try:
                    final_url = new_page_obj.url
                except:
                    final_url = page.url
            else:
                final_url = page.url
            result["final_url"] = final_url

            # Add final landing if not already in chain
            if not navigation_events or navigation_events[-1]["jump_url"] != final_url:
                navigation_events.append({
                    "jump_id": jump_id,
                    "jump_url": final_url,
                    "from_url": navigation_events[-1]["jump_url"] if navigation_events else "",
                    "jump_type": "Final Landing",
                    "status_code": 200,
                    "trigger_type": current_trigger,
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                })

            # Deduplicate and limit jumps
            seen = set()
            valid_jumps = []
            for ev in sorted(navigation_events, key=lambda x: x["jump_id"]):
                url = ev["jump_url"]
                if url and is_valid_url(url) and url not in seen and len(valid_jumps) < CONFIG["max_jump_count"]:
                    # No need to fetch intermediate HTML, keep as is
                    ev["page_html"] = ""  # Not used for classification
                    valid_jumps.append(ev)
                    seen.add(url)

            result["jump_chain"] = valid_jumps

            # Final classification using final page content (safely)
            final_html = ""
            try:
                if new_page_obj and CONFIG["click_config"]["capture_new_window"] and new_page_obj:
                    final_html = new_page_obj.content()[:10000]
                else:
                    final_html = page.content()[:10000]
            except Exception as e:
                logger.warning(f"Failed to get final page content: {e}")
                final_html = ""

            final_jump_info = {
                "jump_url": result["final_url"],
                "page_html": final_html,
                "request_type": response.request.resource_type if response else ""
            }
            result["main_class"], result["sub_class"] = classify_url(final_jump_info)
            result["processing_status"] = "Success"
            logger.info(f"Scan success: {std_url} -> {final_url}")
            break  # success

        except PlaywrightTimeoutError:
            result["error_msg"] = f"Timeout ({CONFIG['timeout']/1000}s)"
            result["error_code"] = "ERR_TIMED_OUT"
            logger.warning(f"Timeout {std_url}, retry {retry+1}")
        except Exception as e:
            error_msg = str(e)
            result["error_msg"] = error_msg
            result["error_code"], _ = extract_error_code(error_msg)
            logger.warning(f"Scan failed {std_url}: {error_msg}, retry {retry+1}")
        finally:
            # Clean up resources
            if page:
                try:
                    page.close()
                except:
                    pass
            if context:
                try:
                    context.close()
                except:
                    pass
            if browser:
                try:
                    browser.close()
                except:
                    pass
            if playwright:
                try:
                    playwright.stop()
                except:
                    pass
            retry += 1

    if result["error_msg"] and not result["main_class"]:
        result["main_class"], result["sub_class"] = classify_url({"error_msg": result["error_msg"]})
    return result

# -------------------------- 6. Excel Export & Statistics --------------------------
def export_to_excel(scan_results: list):
    """Write all scan results to Excel with two sheets (bilingual headers)"""
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = CONFIG["summary_sheet_name"]
    # Bilingual headers for summary
    summary_headers = [
        "No./序号", "Original URL/原始URL", "Standard URL/标准化URL", "Final URL/最终URL",
        "Jump Count/跳转次数", "Full Path/完整路径", "Main Category/主分类", "Sub Category/细分分类",
        "Click Triggered/点击触发", "Click Element Info/点击元素", "Auth Required/需要认证",
        "Status/状态", "Error Message/错误信息", "Error Code/错误码", "Scan Time/扫描时间"
    ]
    summary_ws.append(summary_headers)

    detail_ws = wb.create_sheet(CONFIG["detail_sheet_name"])
    detail_headers = [
        "Original URL/原始URL", "Step/步骤", "URL/跳转URL", "From URL/来源URL",
        "Jump Type/跳转类型", "HTTP Status/状态码", "Trigger Mode/触发方式",
        "Category/分类标签", "Timestamp/时间戳"
    ]
    detail_ws.append(detail_headers)

    for idx, res in enumerate(scan_results, 1):
        # Build summary row
        jump_count = len(res["jump_chain"])
        full_path = " → ".join([j["jump_url"] for j in res["jump_chain"]])
        summary_ws.append([
            idx, res["raw_url"], res["standard_url"], res["final_url"],
            jump_count, full_path, res["main_class"], res["sub_class"],
            "Yes" if res["click_triggered"] else "No", res["click_element_info"],
            "Yes" if res["auth_required"] else "No", res["processing_status"],
            res["error_msg"], res["error_code"], res["scan_time"]
        ])

        # Detail rows
        for step, jump in enumerate(res["jump_chain"], 1):
            detail_ws.append([
                res["raw_url"], step, jump["jump_url"], jump.get("from_url", ""),
                jump.get("jump_type", ""), jump.get("status_code", ""), jump.get("trigger_type", ""),
                classify_url({"jump_url": jump["jump_url"], "page_html": jump.get("page_html", "")})[0],
                jump.get("timestamp", "")
            ])

    # Auto-adjust column widths
    for sheet in [summary_ws, detail_ws]:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 50)
        # Header styling
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    wb.save(CONFIG["output_excel_path"])
    logger.info(f"Excel saved: {CONFIG['output_excel_path']}")

def generate_statistics(scan_results: list):
    """Print stats and generate chart"""
    total = len(scan_results)
    success = sum(1 for r in scan_results if r["processing_status"] == "Success")
    auth_req = sum(1 for r in scan_results if r["auth_required"])

    main_classes = {}
    error_codes = {}
    for r in scan_results:
        cls = r["main_class"]
        main_classes[cls] = main_classes.get(cls, 0) + 1
        if r["error_code"]:
            error_codes[r["error_code"]] = error_codes.get(r["error_code"], 0) + 1

    logger.info("\n" + "="*60)
    logger.info("SCAN STATISTICS | 扫描统计")
    logger.info("="*60)
    logger.info(f"Total URLs: {total}")
    logger.info(f"Success: {success} ({success/total*100:.1f}%)")
    logger.info(f"Failed: {total-success} ({(total-success)/total*100:.1f}%)")
    logger.info(f"Auth Required: {auth_req} ({auth_req/total*100:.1f}%)")
    logger.info("\nMain Category Distribution:")
    for cls, cnt in main_classes.items():
        logger.info(f"  {cls}: {cnt} ({cnt/total*100:.1f}%)")
    if error_codes:
        logger.info("\nError Code Distribution:")
        for code, cnt in error_codes.items():
            logger.info(f"  {code}: {cnt}")

    # Chart
    if main_classes:
        plt.rcParams["font.sans-serif"] = ["DejaVu Sans"]
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
        labels = list(main_classes.keys())
        sizes = list(main_classes.values())
        ax1.pie(sizes, labels=labels, autopct="%1.1f%%", startangle=90)
        ax1.set_title("Main Category Distribution")
        if error_codes:
            codes = list(error_codes.keys())
            counts = list(error_codes.values())
            ax2.bar(codes, counts)
            ax2.set_title("Error Code Distribution")
            ax2.tick_params(axis="x", rotation=45)
        plt.tight_layout()
        chart_path = f"url_scan_chart_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        plt.savefig(chart_path, dpi=300, bbox_inches="tight")
        logger.info(f"Chart saved: {chart_path}")

def get_processed_urls(output_path: str) -> set:
    """Read already processed URLs from output Excel (breakpoint resume)"""
    if not os.path.exists(output_path):
        return set()
    try:
        wb = load_workbook(output_path, read_only=True)
        ws = wb[CONFIG["summary_sheet_name"]]
        processed = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            # column 2 is "Original URL"
            raw_url = row[1]
            status = row[11]  # Status column
            if raw_url and status == "Success":
                processed.add(raw_url)
        wb.close()
        return processed
    except Exception as e:
        logger.warning(f"Could not read processed URLs: {e}")
        return set()

# -------------------------- 7. Main Function --------------------------
def main():
    parse_command_line_args()
    if not os.path.exists(CONFIG["input_excel_path"]):
        logger.error(f"Input file not found: {CONFIG['input_excel_path']}")
        return

    # Load URLs from Excel
    try:
        wb = load_workbook(CONFIG["input_excel_path"], read_only=True)
        ws = wb[CONFIG["input_sheet_name"]]
        url_column = CONFIG["url_column"]
        col_idx = ord(url_column.upper()) - 64 if len(url_column) == 1 else None
        if not col_idx:
            logger.error("Invalid url_column config, use 'A' as fallback")
            col_idx = 1
        url_list = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and len(row) >= col_idx:
                val = row[col_idx-1]
                if val and isinstance(val, str) and val.strip():
                    url_list.append(val.strip())
        wb.close()
    except Exception as e:
        logger.error(f"Failed to read input Excel: {e}")
        return

    # Deduplicate
    url_list = list(set(url_list))
    # Breakpoint resume
    processed = get_processed_urls(CONFIG["output_excel_path"])
    pending = [u for u in url_list if u not in processed]
    logger.info(f"Total URLs: {len(url_list)}, Already processed: {len(processed)}, Pending: {len(pending)}")
    if not pending:
        logger.info("All URLs already processed. Exiting.")
        return

    # Multi-thread scan
    results = []
    with ThreadPoolExecutor(max_workers=CONFIG["max_workers"]) as executor:
        future_to_url = {executor.submit(scan_single_url, url): url for url in pending}
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            try:
                res = future.result()
                results.append(res)
                logger.info(f"Completed: {url} -> {res['main_class']}")
            except Exception as e:
                logger.error(f"Task failed for {url}: {traceback.format_exc()}")
                results.append({
                    "raw_url": url, "standard_url": "", "jump_chain": [],
                    "main_class": "Error", "sub_class": "Task Exception",
                    "error_msg": str(e), "error_code": "TASK_ERROR",
                    "processing_status": "Failed", "scan_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                })

    # Export and statistics
    export_to_excel(results)
    generate_statistics(results)
    logger.info("All tasks completed.")

if __name__ == "__main__":
    main()